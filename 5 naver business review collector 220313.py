from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import time
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import openpyxl as oxl
import requests
import json
import os

real_df = []
real_df1 = []
real_df2 = []
names = []
j = 0

curr_date = datetime.strftime(datetime.now(), "%y_%m_%d")
#구글지도를 켜고 마우스 오른쪽버튼을 누르면 위도와 경도가 나옴
lati_logi = input('검색할 지역의 위도와 경도를 입력하세요 : ')
url1 = lati_logi.split(',')
search_area = input('검색할 지역을 입력하세요 : ')
search_business = input('검색할 사업을 입력하세요 : ')
search_code = input('검색할 사업 코드를 입력하세요(ex: 1: restaurant, 2: hospital, 3: place), 기타는 입력하세요. : ')
if search_code == '1':
    search_code = 'restaurant'
elif search_code == '2':
    search_code = 'hospital'
elif search_code == '3':
    search_code = 'place'

browser = webdriver.Chrome(ChromeDriverManager().install())
wait = WebDriverWait(browser, 5)#브라우저 실행시 최대 5초간 대기
for i in range(1):
  url = 'https://map.naver.com/v5/api/search?query=' + search_business + '&type=place&searchCoord=' + str(url1[1]) + ';' + str(url1[0]) + '&page=' + str(i) + '&displayCount=100'
  payload = json.dumps({
      "location": "https://naveropenapi.apigw.ntruss.com/map-static/v2/raster/ncpclientid=ku93d0gqaq",
      # "X-NCP-APIGW-API-KEY-ID": "ku93d0gqaq",
      "X-NCP-APIGW-API-KEY": "vsdzhy0I0npUHmTn8mj1cKSKYVudzOryL7TYN9s7",
      "argument": {
          "type": "1"
      }
  })
  headers = {
    'authority': 'map.naver.com',
    'method': 'GET',
    'scheme': 'https',
    'accept': 'application/json, text/plain, */*',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'ko-KR,ko;q=0.8,en-US;q=0.6,en;q=0.4',
    'cache-control': 'no-cache',
    }
  response = requests.get(url, headers=headers, data=payload)
  text = response.text
  data = json.loads(text)
#네이버지도상의 선택된 비즈니스에 대해 이름과 주소, ID를 추출함
  for list in data['result']['place']['list']:
    name = list['name']
    address = list['address']
    id = list['id']
    print(j, name, address, id)
    j += 1
    if [name, address, id] in names:#이름, 주소, id가 있으면 진행하지 않음
        continue
#해당업체의 리뷰주소를 읽어오기
    url = 'https://pcmap.place.naver.com/' + search_code + '/' + id + '/review/visitor'#?reviewItem=0'
    browser.get(url)
    time.sleep(1.5)
    html = browser.page_source
    soup = BeautifulSoup(html, 'html.parser')

#업체별 장점파악하기
    advantage = []
    if soup.find('span', 'Nqp-s') == None:
        pass
    else:
        try:
            advantage_button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section._11ptV > div > div > div._10UcK > a')))
        except:
            advantage_reviews = soup.find_all('span', '_1lntw')
            advantage_nums = soup.find_all('span', 'Nqp-s')
            for advantage_all in advantage_alls:
                advantage_two = advantage_all.text.split('이 키워드를 선택한 인원')
                advantage.append([advantage_two[0].strip('"'), int(advantage_two[1])])
        else:
            advantage_button.click()
            time.sleep(1)
            html = browser.page_source
            soup = BeautifulSoup(html, 'html.parser')
            advantage_alls = soup.find_all('div', '_3ZEZK')
            i = 0
            for advantage_all in advantage_alls:
                advantage_two = advantage_all.text.split('이 키워드를 선택한 인원')
                advantage.append([advantage_two[i].strip('"'), int(advantage_two[i+1])])
    columns1 = ['장점', '투표수']
    real_df1 = pd.DataFrame(advantage, columns=columns1)

#리뷰 총회수 구하기
    if soup.find('span', 'place_section_count') == None:
        continue
    else:
        review_sum = soup.find('span', 'place_section_count').text#전체 리뷰수를 읽어옴
        review_sum = int(review_sum.replace(',', ''))
        review_num = review_sum / 10
    names.append([name, address, id, review_sum])#이름, 주소, ID, 총리뷰수를 저장
#리뷰수를 10으로 나누어 1보다 작으면 바로 읽기, 5보다 크면 5회 리뷰 읽기, 1보다 크고 5보다 작으면 해당 횟수만큼 읽기
    if review_num <= 1:
        pass
    elif review_num > 5:
        review_num = 5
        try:
            body = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section.cXO6M > div._2kAri > a > svg')))
            for i in range(int(review_num)):
                    body.click()#더보기 버튼을 실행
                    time.sleep(1)
        except:
            print(name, address)
            continue
    else:
        try:
            body = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#app-root > div > div > div.place_detail_wrapper > div:nth-child(5) > div:nth-child(4) > div.place_section.cXO6M > div._2kAri > a > svg')))
            for i in range(int(review_num)):
                    body.click()
                    time.sleep(1)
        except:
            print(name, address, id)
            continue
#리뷰 읽어오기
    html = browser.page_source
    soup = BeautifulSoup(html, 'html.parser')
    lists = soup.find_all('li', '_3FaRE')
    list_data = []
    for list in lists:
        if list.find('span', 'WoYOw') == None:#리뷰 읽기
            continue
        else:
            review = list.find('span', 'WoYOw').text
            review = ILLEGAL_CHARACTERS_RE.sub(r'', review)
        visites = list.find_all('div', '_3-LAD')
        for visite in visites:
            dates = visite.find_all('span', '_1fvo3')
            for date in dates:
                if '최근 방문일' in date.text:
                    visite_date = date.find_all('span', 'place_blind')#방문일자
                    visite_date = visite_date[1].text
                elif '번째 방문' in date.text:
                    num = int(date.text.strip('번째 방문'))#방문횟수
        list_data.append([review,visite_date, num])#리뷰 저장
    columns = ['리뷰', '방문날짜', '방문횟수']
    real_df = pd.DataFrame(list_data, columns=columns).sort_values(by='방문날짜', ascending=False)

#파일 쉬트에 저장, 파일이 있을 경우와 없을 경우 구분
    if not os.path.exists('5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx'):
        with pd.ExcelWriter('./5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx', mode='w', engine='openpyxl') as writer:
            real_df1.to_excel(writer, sheet_name=f"{name}{','}{review_sum}", startrow=1)
            real_df.to_excel(writer, sheet_name=f"{name}{','}{review_sum}", startrow=15)
    else:
        with pd.ExcelWriter('./5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            real_df1.to_excel(writer, sheet_name=f"{name}{','}{review_sum}", startrow=1)
            real_df.to_excel(writer, sheet_name=f"{name}{','}{review_sum}", startrow=15)

#업체별 상호, 주소, ID, 리뷰슈를 하나의 쉬트에 정리
columns2 = ['상호명', '주소', 'ID', '리뷰수']
real_df2 = pd.DataFrame(names, columns=columns2).sort_values(by='리뷰수', ascending=False)
with pd.ExcelWriter('./5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx', mode='a', engine='openpyxl') as writer:
    real_df2.to_excel(writer, sheet_name=f"{search_area}")
#쉬트 정렬하기
numbers = []
wb = oxl.load_workbook('./5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx')
ws_names = wb.sheetnames
for ws_name in ws_names:
    name = ws_name.split(',')
    if len(name) == 1:
        break
    else:
        numbers.append(int(name[1]))
numbers.sort(reverse=True)
for i_i in range(len(numbers)):
    ws_names = wb.sheetnames
    for j_j in range(len(ws_names)-1):
        if numbers[i_i] == int(ws_names[j_j].split(',')[1]):
            ws = wb[ws_names[j_j]]
            wb.move_sheet(ws, i_i-j_j)
            break
wb.move_sheet(ws_names[-1], -len(ws_names) + 1)
wb.save('./5 ' + search_area + '지역 ' + search_business + ' review-collector ' + curr_date + '.xlsx')