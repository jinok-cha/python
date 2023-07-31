from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from datetime import datetime
import pandas as pd
import openpyxl
import requests
import json
import time
import os

j = 0
dong_count = 0
# list_data = []
data = []
sheet_list = []
macket_rate = input('시장환원률을 입력하세요 :')
curr_date = datetime.strftime(datetime.now(), "%y_%m_%d")
filename = '/Volumes/ext1/차진옥/python/pusan core dong 220314.xlsx'
book = openpyxl.load_workbook(filename)
sheet = book.worksheets[0]
excel_url = '6 부산지역 주요 동 네이버부동산 매물조회 ' + curr_date + '.xlsx'
try:
    book1 = openpyxl.load_workbook(excel_url)
except:
    sheet_count = 0
    pass
else:
    book1 = openpyxl.load_workbook(excel_url)
    last_sheet = book1.worksheets
    for row in sheet.rows:
        sheet_list.append(row[0].value)
    sheet_count = sheet_list.index(last_sheet[-1].title)+1
sheet_counts = 0


for row in sheet.rows:
    if sheet_count == sheet_counts:
        pass
    else:
        sheet_counts += 1
        continue
    list_data = []
    dong_code = row[1].value
    print(row[0].value)
    if dong_count % 4 == 0:
        time.sleep(30)
    for i in range(5):
        url = 'https://new.land.naver.com/api/articles?cortarNo=' + str(dong_code) + '&order=rank&realEstateType=GM&priceType=RETAIL&page=' + str(i+1)
        payload={}
        headers = {
            'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE2NDcxNzg2NTQsImV4cCI6MTY0NzE4OTQ1NH0.jxNhCqLWDDvHPvF8KbqULExCnEujbv67wGpws4ibK-U',
            'User-Agent': "PostmanRuntime/7.20.0",
            'Accept': "*/*",
            'Cache-Control': "no-cache",
            'Postman-Token': "adbba748-cb85-4fb4-8f6a-4be441f19cc3",
            'Host': "m.land.naver.com",
            'Accept-Encoding': "gzip, deflate",
            'Connection': "keep-alive",
            'cache-control': "no-cache"
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        text = response.text
        data = json.loads(text)
        if 'articleList' in data:
            for list in data['articleList']:
                id = list['articleNo']
                url = 'https://new.land.naver.com/api/articles/' + str(id) + '?complexNo='
                payload = {}
                headers = {    'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE2NDcxNzg2NTQsImV4cCI6MTY0NzE4OTQ1NH0.jxNhCqLWDDvHPvF8KbqULExCnEujbv67wGpws4ibK-U',
            'User-Agent': "PostmanRuntime/7.20.0",
            'Accept': "*/*",
            'Cache-Control': "no-cache",
            'Postman-Token': "adbba748-cb85-4fb4-8f6a-4be441f19cc3",
            'Host': "m.land.naver.com",
            'Accept-Encoding': "gzip, deflate",
            'Connection': "keep-alive",
            'cache-control': "no-cache"
            }
                response = requests.request("GET", url, headers=headers, data=payload)
                text = response.text
                data1 = json.loads(text)

                articleNo = data1['articleDetail']['articleNo']
                try:
                    noneAptBuildingName = data1['articleDetail']['noneAptBuildingName']
                except:
                    noneAptBuildingName = " "
                    pass
                try:
                    articleConfirmYMD = data1['articleDetail']['articleConfirmYMD']
                except:
                    articleConfirmYMD = " "
                    pass
                try:
                    lawUsage = data1['articleDetail']['lawUsage']
                except:
                    lawUsage = " "
                    pass
                try:
                    floorInfo = data1['articleAddition']['floorInfo']
                except:
                    floorInfo = " "
                    pass
                try:
                    area1 = data1['articleAddition']['area1']
                except:
                    area1 = 0.0
                    pass
                try:
                    area2 = data1['articleAddition']['area2']
                except:
                    area2 = 0.0
                    pass
                try:
                    cpPcArticleUrl = data1['articleAddition']['cpPcArticleUrl']
                except:
                    cpPcArticleUrl = " "
                    pass
                try:
                    directionTypeName = data1['articleFacility']['directionTypeName']
                except:
                    directionTypeName = " "
                    pass
                try:
                    usageAreaTypeName = data1['articleFacility']['usageAreaTypeName']
                except:
                    usageAreaTypeName = " "
                    pass
                try:
                    buildingUseAprvYmd = data1['articleFacility']['buildingUseAprvYmd']
                except:
                    buildingUseAprvYmd = " "
                    pass
                try:
                    dealPrice = data1['articlePrice']['dealPrice']
                except:
                    dealPrice = 0
                    pass
                try:
                    allWarrantPrice = data1['articlePrice']['allWarrantPrice']
                except:
                    allWarrantPrice = 0
                    pass
                try:
                    allRentPrice = data1['articlePrice']['allRentPrice']
                except:
                    allRentPrice = 0
                    pass
                value = (allRentPrice * 12) / float(macket_rate) + allWarrantPrice
                value1 = (allRentPrice * 12) / (float(macket_rate) - 0.005) + allWarrantPrice
                if dealPrice - value < 0:
                    value_check = 1
                elif dealPrice - value1 < 0:
                    value = value1
                    value_check = 2
                else:
                    value_check = '-'
                try:
                    priceBySpace = data1['articlePrice']['priceBySpace']
                except:
                    priceBySpace = 0
                    pass
                try:
                    realtorName = data1['articleRealtor']['realtorName']
                except:
                    realtorName = " "
                    pass
                try:
                    representativeName = data1['articleRealtor']['representativeName']
                except:
                    representativeName = " "
                    pass
                try:
                    cellPhoneNo = data1['articleRealtor']['cellPhoneNo']
                except:
                    cellPhoneNo = " "
                    pass
                try:
                    etcParkInfo = data1['articleBuildingRegister']['etcParkInfo'].strip('(').strip(')')
                except:
                    etcParkInfo = " "
                    pass
                try:
                    bcRat = data1['articleBuildingRegister']['bcRat']
                except:
                    bcRat = 0.0
                    pass
                try:
                    vlRat = data1['articleBuildingRegister']['vlRat']
                except:
                    vlRat = 0.0
                    pass
                try:
                    newPlatPlc = data1['articleBuildingRegister']['newPlatPlc']
                except:
                    newPlatPlc = 0.0
                    pass
                j += 1
                list_data.append([row[0].value, articleNo, dealPrice, priceBySpace, noneAptBuildingName, area1, area2, bcRat, vlRat, usageAreaTypeName, lawUsage, floorInfo, directionTypeName, etcParkInfo, \
                                  buildingUseAprvYmd, allWarrantPrice, allRentPrice, int(value), value_check, newPlatPlc, articleConfirmYMD, realtorName, representativeName, cellPhoneNo, cpPcArticleUrl])
        # if j % 300 == 0:
        #     time_count = 20
        # else:
        #     time_count = 5
        # time.sleep(time_count)
# 파일 쉬트에 저장, 파일이 있을 경우와 없을 경우 구분
    columns = ['지역', '매물번호', '매매가격', '평단가', '건물명', '대지면적', '연면적', '건폐율', '용적률', '토지용도', '건물용도', '층수', '방향', '주차', '사용승인일자', '보증금', '월세', '부동산가치', '확인', '물건주소', '확인일자', '중개사무소', '중개인', '전화', '블로그']
    real_df = pd.DataFrame(list_data, columns=columns)
    if not os.path.exists('./6 부산지역 주요 동 네이버부동산 매물조회 ' + curr_date + '.xlsx'):
        with pd.ExcelWriter('./6 부산지역 주요 동 네이버부동산 매물조회 ' + curr_date + '.xlsx', mode='w', engine='openpyxl') as writer:
            real_df.to_excel(writer, sheet_name=f"{row[0].value}")
    else:
        with pd.ExcelWriter('./6 부산지역 주요 동 네이버부동산 매물조회 ' + curr_date + '.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            real_df.to_excel(writer, sheet_name=f"{row[0].value}")
    time.sleep(40)

#모든 쉬트를 읽어와서 하나로 합치기
columns = ['지역', '매물번호', '매매가격', '평단가', '건물명', '대지면적', '연면적', '건폐율', '용적률', '토지용도', '건물용도', '층수', '방향', '주차', '사용승인일자', '보증금', '월세', '부동산가치', '확인', '확인일자', '중개사무소', '중개인', '전화', '블로그']
df_all = pd.read_excel(excel_url, sheet_name = None)
concatted_df = pd.concat(df_all, ignore_index=False)
real_df2 = pd.DataFrame(concatted_df, columns=columns).sort_values(by='대지면적', ascending=False)
with pd.ExcelWriter('./6 부산지역 주요 동 네이버부동산 매물조회 ' + curr_date + '.xlsx', mode='a', engine='openpyxl') as writer:
    real_df2.to_excel(writer, sheet_name=f"{curr_date}")