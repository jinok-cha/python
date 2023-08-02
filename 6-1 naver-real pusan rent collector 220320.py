from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
import openpyxl
import requests
import json
import time
import os

j = 0
list_data = []
data = []
sheet_list = []
dong_count = 0

curr_date = datetime.strftime(datetime.now(), "%y_%m_%d")
filename = '/Volumes/ext1/jinok/python/pusan core dong 220314.xlsx'
book = openpyxl.load_workbook(filename)
sheet = book.worksheets[0]
excel_url = '6-1 부산지역 주요 동 네이버부동산 월세 매물조회 ' + curr_date + '.xlsx'
browser = webdriver.Chrome(ChromeDriverManager().install())
wait = WebDriverWait(browser, 5)  # 브라우저 실행시 최대 5초간 대기
try:
    book1 = openpyxl.load_workbook(excel_url)
except:
    sheet_count = 0
    pass
else:
    book1 = openpyxl.load_workbook(excel_url)
    last_sheet = book1.worksheets
    for row in sheet.rows:
        sheet_list.append(row[0].value)
    sheet_count = sheet_list.index(last_sheet[-1].title)+1
sheet_counts = 0

for row in sheet.rows:
    if sheet_count == sheet_counts:
        pass
    else:
        sheet_counts += 1
        continue
    list_data = []
    lati_code = row[2].value
    long_code = row[3].value
    print(row[0].value)
    if dong_count % 4 == 0:
        time.sleep(10)
    leftLon = round(float(long_code) - 0.0068665, 7)
    rightLon = round(float(long_code) + 0.0068665, 7)
    topLat = round(float(lati_code) + 0.0031733, 7)
    bottomLat = round(float(lati_code) - 0.0031733, 7)

    for i in range(10):
        url = 'https://new.land.naver.com/api/articles?zoom=17&&realEstateType=SG%3ASMS&tradeType=&priceType=RETAIL&&articleState&page=' + str(i+1) + '&leftLon=' + str(leftLon) + '&rightLon=' + str(rightLon) + '&topLat=' + str(topLat) + '&bottomLat=' + str(bottomLat) + '&order=rank'
        payload={}
        headers = {
            'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE2NDc0Mzc1MjksImV4cCI6MTY0NzQ0ODMyOX0.BFkee1Qrgl_X5xEQ2iNrFfIu4EZ6VI2scTiW0Q8Rxfo',
            'Accept': "*/*",
            'Accept-Encoding': "gzip, deflate, br",
            'Accept-Language': "ko-KR, ko;q=0.9, en-US;q=0.8, en;q=0.7",
            'Cache-Control': "no-cache",
            'Postman-Token': "adbba748-cb85-4fb4-8f6a-4be441f19cc3",
            'Host': "m.land.naver.com",
            'Connection': "keep-alive",
            'cache-control': "no-cache",
            'User-Agent': "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.104 Whale/3.13.131.36 Safari/537.36"
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        text = response.text
        data = json.loads(text)
        if 'articleList' in data:
            for list in data['articleList']:
                id = list['articleNo']
                url = 'https://new.land.naver.com/api/articles/' + str(id) + '?complexNo='
                payload = {}
                headers = {
            'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE2NDczMjA4NDAsImV4cCI6MTY0NzMzMTY0MH0.CVMz-aQokatAj0AXsYi6maU5zvzYTBurOQfhW7XUpnU',
            'User-Agent': "PostmanRuntime/7.20.0",
            'Accept': "*/*",
            'Cache-Control': "no-cache",
            'Postman-Token': "adbba748-cb85-4fb4-8f6a-4be441f19cc3",
            'Host': "m.land.naver.com",
            'Accept-Encoding': "gzip, deflate",
            'Connection': "keep-alive",
            'cache-control': "no-cache"
            }
                response = requests.request("GET", url, headers=headers, data=payload)
                text = response.text
                data1 = json.loads(text)
                articleNo = data1['articleDetail']['articleNo']
                try:
                    articleName = data1['articleDetail']['articleName']
                except:
                    articleName = " "
                    pass
                try:
                    articleConfirmYMD = data1['articleDetail']['articleConfirmYMD']
                except:
                    articleConfirmYMD = " "
                    pass
                try:
                    lawUsage = data1['articleDetail']['lawUsage']
                except:
                    lawUsage = " "
                    pass
                try:
                    floorInfo = data1['articleAddition']['floorInfo']
                    floorInfo = floorInfo.split('/')
                    if 'B' in floorInfo[0]:
                        floorInfo[0] = floorInfo[0].replace('B', '-')
                    if '전체층' in floorInfo[0]:
                        floorInfo[0] = 999
                except:
                    floorInfo[0] = 0
                    floorInfo[1] = 0
                    pass
                try:
                    exclusiveSpace = data1['articleSpace']['exclusiveSpace']
                except:
                    exclusiveSpace = 0.0
                    pass
                try:
                    supplySpace = data1['articleSpace']['supplySpace']
                except:
                    supplySpace = 0.0
                    pass
                try:
                    cpPcArticleUrl = data1['articleAddition']['cpPcArticleUrl']
                except:
                    cpPcArticleUrl = " "
                    pass
                try:
                    latitude = data1['articleAddition']['latitude']
                except:
                    latitude = " "
                    pass
                try:
                    longitude = data1['articleAddition']['longitude']
                except:
                    longitude = " "
                    pass
                try:
                    directionTypeName = data1['articleFacility']['directionTypeName']
                except:
                    directionTypeName = " "
                    pass
                try:
                    buildingUseAprvYmd = data1['articleFacility']['buildingUseAprvYmd']
                except:
                    buildingUseAprvYmd = " "
                    pass
                try:
                    warrantPrice = data1['articlePrice']['warrantPrice']
                except:
                    warrantPrice = 0
                    pass
                try:
                    rentPrice = data1['articlePrice']['rentPrice']
                except:
                    allRentPrice = 0
                    pass
                try:
                    priceBySpace = int((((warrantPrice * 0.06) /12 + rentPrice) / exclusiveSpace) * 3.3 * 10000)
                except:
                    priceBySpace = 0
                    pass
                try:
                    realtorName = data1['articleRealtor']['realtorName']
                except:
                    realtorName = " "
                    pass
                try:
                    representativeName = data1['articleRealtor']['representativeName']
                except:
                    representativeName = " "
                    pass
                try:
                    cellPhoneNo = data1['articleRealtor']['cellPhoneNo']
                except:
                    cellPhoneNo = " "
                    pass
                try:
                    etcParkInfo = data1['articleBuildingRegister']['etcParkInfo'].strip('(').strip(')')
                except:
                    etcParkInfo = " "
                    pass
                try:
                    newPlatPlc = data1['articleBuildingRegister']['newPlatPlc']
                except:
                    newPlatPlc = ''
                    pass
                try:
                    pnu = data1['articleBuildingRegister']['pnu']
                except:
                    pnu = ''
                    pass
    #개별공시지가 불러오기
                url = 'https://data.disco.re/home/land_price/?pnu=' + pnu
                browser.get(url)
                time.sleep(0.5)
                html = browser.page_source
                soup = BeautifulSoup(html, 'html.parser')
                try:
                    land_prices = soup.text.strip('[').strip(']').split(',')
                    land_price_year = land_prices[-3].strip(' [')
                except:
                    land_prices[-1] = 0
                    land_price_year = 0
                    pass
                j += 1
                list_data.append([articleNo, articleName, newPlatPlc, exclusiveSpace, supplySpace, int(floorInfo[0]), int(floorInfo[1]), warrantPrice, rentPrice, priceBySpace, int(land_prices[-1]), lawUsage, directionTypeName, etcParkInfo, \
                                  buildingUseAprvYmd, latitude, longitude, articleConfirmYMD, realtorName, representativeName, cellPhoneNo, cpPcArticleUrl])
        time.sleep(5)
    columns = ['매물번호', '건물명', '물건주소', '전용면적', '계약면적', '임대층', '총층수', '보증금', '월세', '평단가', '공시지가', '건물용도', '방향', '주차', '사용승인일자', '위도', '경도', '확인일자', '중개사무소', '중개인', '전화', '블로그']
    real_df = pd.DataFrame(list_data, columns=columns).sort_values(by='전용면적', ascending=False)
    if not os.path.exists('./6-1 부산지역 주요 동 네이버부동산 월세 매물조회 ' + curr_date + '.xlsx'):
        with pd.ExcelWriter('./6-1 부산지역 주요 동 네이버부동산 월세 매물조회 ' + curr_date + '.xlsx', mode='w', engine='openpyxl') as writer:
            real_df.to_excel(writer, sheet_name=f"{row[0].value}")
    else:
        with pd.ExcelWriter('./6-1 부산지역 주요 동 네이버부동산 월세 매물조회 ' + curr_date + '.xlsx', mode='a', engine='openpyxl') as writer:
            real_df.to_excel(writer, sheet_name=f"{row[0].value}")

#모든 쉬트를 읽어와서 하나로 합치기
columns = ['매물번호', '건물명', '물건주소', '전용면적', '계약면적', '임대층', '총층수', '보증금', '월세', '평단가', '공시지가', '건물용도', '방향', '주차', '사용승인일자', '위도', '경도', '확인일자', '중개사무소', '중개인', '전화', '블로그']
df_all = pd.read_excel(excel_url, sheet_name = None)
concatted_df = pd.concat(df_all, ignore_index=False)
real_df2 = pd.DataFrame(concatted_df, columns=columns).sort_values(by='전용면적', ascending=False)
with pd.ExcelWriter('./6-1 부산지역 주요 동 네이버부동산 월세 매물조회 ' + curr_date + '.xlsx', mode='a', engine='openpyxl') as writer:
    real_df2.to_excel(writer, sheet_name=f"{curr_date}")